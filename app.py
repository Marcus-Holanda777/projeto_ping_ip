import subprocess
import platform
import openpyxl as wb
import unicodedata
import re
import sys
import os
from tela import menu
from time import sleep
from collections import deque
from datetime import datetime

ping_ok = re.compile(r'Recebidos = 1')
ping_inacessivel = re.compile(r'(Host|Rede) de destino inacess.vel')

TMP_EXEC = datetime.now().strftime('%Y%m%d_%Hh%Mm%Ss')
EXCEL_BASE = 'IP_lojas.xlsx'
LINHA_INICIAL = 9
EXCEL_BASE_SAIDA = f'IP_lojas_RESULTADO_{TMP_EXEC}.xlsx'

logs = deque(maxlen=2)


def remover_acento(text):
    n = unicodedata.normalize('NFD', text)
    nn = ''.join([x for x in n if not unicodedata.combining(x)])

    return unicodedata.normalize('NFC', nn)


def gera_log(text):
    text = text.replace('\n', ':')
    with open('log.txt', 'a', newline='') as arq:
        arq.write(text + '\n')


def ping(endereco_ip):
    ping_str = "-n 1" if platform.system().lower() == "windows" else "-c 1"
    args = f'ping {ping_str} {endereco_ip}'
    need_sh = False if platform.system().lower() == "windows" else True

    saida = subprocess.Popen(
        args, shell=need_sh, stdout=subprocess.PIPE, text=True)
    return saida


def main():
    # 1 - carregar excel
    pasta = wb.load_workbook(EXCEL_BASE)
    plan = pasta.active
    ultima_linha = plan.max_row

    pasta_root = 'LOG_PING'

    if not os.path.isdir(pasta_root):
        os.mkdir(pasta_root)

    saida_padrao = os.path.join(pasta_root, EXCEL_BASE_SAIDA)

    for row in range(LINHA_INICIAL, ultima_linha + 1):
        msg = f'Loja: {plan.cell(row, 1).value}. IP: {plan.cell(row, 3).value}'

        saida = ping(plan.cell(row, 3).value)
        ver = ''.join([line for line in saida.stdout])

        ver = remover_acento(ver)

        if ping_inacessivel.search(ver):
            comp = '[INACESSIVEL]'.center(20)
            plan.cell(row, 5, '3 - INACESSIVEL')
            gera_log(f'Filial: {plan.cell(row, 1).value}, {ver}')

        else:
            if ping_ok.search(ver):
                comp = '[RECEBIDO]'.center(20)

                plan.cell(row, 5, '1 - RECEBIDO')
            else:
                comp = '[TEMPO ESGOTADO]'.center(20)
                plan.cell(row, 5, '2 - TEMPO ESGOTADO')

                gera_log(f'Filial: {plan.cell(row, 1).value}, {ver}')
        msg = msg + ',' + comp

        logs.appendleft(msg)

        msg_log = ''
        for log in logs:
            msg_log += log + ' '

        sys.stdout.write('\r' + msg_log)
        sys.stdout.flush()

    pasta.save(saida_padrao)
    pasta.close()


menu('MENU / VERIFICA PING  IP')
print('')
print('===>')
inicia = input('Iniciar: (s/n): ')
while inicia.lower() != 'n' and inicia.lower() != 's':
    inicia = input('Iniciar: (s/n): ')

if inicia.lower() == 'n':
    sys.exit()

elif inicia.lower() == 's':
    if not os.path.isfile(EXCEL_BASE):
        print(f'\t\t -->> [ERRO] - Arquivo {EXCEL_BASE}, não existe !')
        sleep(5)
        sys.exit()
    else:
        print('')
        print('\t\t -->> Filial ... ping .. inicio !')
        main()
